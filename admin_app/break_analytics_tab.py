# admin_app/break_analytics_tab.py
"""
Вкладка аналитики системы перерывов

Включает:
- Dashboard (кто в перерыве, кто превышает лимит, топ нарушителей)
- Фильтры (дата, сотрудник, группа, тип нарушения)
- Таблица нарушений
- Статистика
- Экспорт в Excel
- Отчёты и графики
"""
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox,
    QDateEdit, QComboBox, QLineEdit, QSplitter, QFrame,
    QMessageBox, QFileDialog, QTabWidget
)
from PyQt5.QtCore import Qt, QDate, QTimer
from PyQt5.QtGui import QFont, QColor
from datetime import datetime, timedelta, date
import logging
import sys
from pathlib import Path

# Добавляем путь к shared модулям
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
from shared.time_utils import format_datetime_moscow, format_time_moscow

logger = logging.getLogger(__name__)


class BreakAnalyticsTab(QWidget):
    """Вкладка аналитики перерывов"""
    
    def __init__(self, break_manager, parent=None):
        super().__init__(parent)
        self.break_mgr = break_manager
        self.current_violations = []
        self.dashboard_active_breaks_data = []  # Данные для клика
        self.dashboard_over_limit_data = []     # Данные для клика
        self._setup_ui()
        
        # Автообновление Dashboard каждые 30 секунд
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.refresh_dashboard)
        self.timer.start(30000)
        
        # Начальная загрузка
        self.refresh_dashboard()
        self.apply_filters()
    
    def _setup_ui(self):
        """Создаёт интерфейс"""
        layout = QVBoxLayout(self)
        
        # Вкладки
        tabs = QTabWidget()
        
        # Вкладка 1: Dashboard + Нарушения
        main_tab = self._build_main_tab()
        tabs.addTab(main_tab, "📊 Dashboard и Нарушения")
        
        # Вкладка 2: Отчёты
        reports_tab = self._build_reports_tab()
        tabs.addTab(reports_tab, "📈 Отчёты и Графики")
        
        layout.addWidget(tabs)
    
    def _build_main_tab(self):
        """Основная вкладка с Dashboard и нарушениями"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Dashboard
        dashboard = self._build_dashboard()
        layout.addWidget(dashboard)
        
        # Разделитель
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        layout.addWidget(line)
        
        # Фильтры
        filters = self._build_filters()
        layout.addWidget(filters)
        
        # Статистика
        stats = self._build_stats()
        layout.addWidget(stats)
        
        # Таблица нарушений
        table_group = QGroupBox("Нарушения")
        table_layout = QVBoxLayout()
        
        # Кнопки над таблицей
        buttons_layout = QHBoxLayout()
        
        btn_refresh = QPushButton("🔄 Обновить")
        btn_refresh.clicked.connect(self.apply_filters)
        buttons_layout.addWidget(btn_refresh)
        
        btn_export = QPushButton("📥 Экспорт в Excel")
        btn_export.clicked.connect(self.export_to_excel)
        buttons_layout.addWidget(btn_export)
        
        buttons_layout.addStretch()
        table_layout.addLayout(buttons_layout)
        
        # Таблица
        self.violations_table = QTableWidget()
        self.violations_table.setColumnCount(7)
        self.violations_table.setHorizontalHeaderLabels([
            "Дата/Время", "Сотрудник", "Тип", "Тип нарушения", "Детали", "Критичность", "Статус"
        ])
        self.violations_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.violations_table.horizontalHeader().setStretchLastSection(True)
        self.violations_table.setAlternatingRowColors(True)
        table_layout.addWidget(self.violations_table)
        
        table_group.setLayout(table_layout)
        layout.addWidget(table_group)
        
        return widget
    
    def _build_dashboard(self):
        """Создаёт Dashboard"""
        group = QGroupBox("📊 DASHBOARD - Реалтайм мониторинг")
        main_layout = QVBoxLayout()
        
        # Кнопка обновить
        refresh_btn = QPushButton("🔄 Обновить")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-weight: bold;
                padding: 8px 16px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        refresh_btn.clicked.connect(self.refresh_dashboard)
        
        # Добавляем кнопку справа
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(refresh_btn)
        main_layout.addLayout(btn_layout)
        
        # Карточки
        cards_layout = QHBoxLayout()
        
        # Карточка 1: Кто сейчас в перерыве
        card1 = self._create_dashboard_card(
            "👥 Сейчас в перерыве",
            "0 человек",
            "#3498db",
            "active_breaks"
        )
        cards_layout.addWidget(card1)
        
        # Карточка 2: Превышают лимит
        card2 = self._create_dashboard_card(
            "⚠️ Превышают лимит",
            "0 человек",
            "#e74c3c",
            "over_limit"
        )
        cards_layout.addWidget(card2)
        
        # Карточка 3: Нарушений сегодня
        card3 = self._create_dashboard_card(
            "📉 Нарушений сегодня",
            "0",
            "#f39c12",
            "today_violations"
        )
        cards_layout.addWidget(card3)
        
        # Карточка 4: Топ нарушитель
        card4 = self._create_dashboard_card(
            "🏆 Топ нарушитель",
            "Нет данных",
            "#9b59b6",
            "top_violator"
        )
        cards_layout.addWidget(card4)
        
        main_layout.addLayout(cards_layout)
        group.setLayout(main_layout)
        return group
    
    def _create_dashboard_card(self, title, value, color, card_id):
        """Создаёт карточку Dashboard"""
        card = QFrame()
        card.setFrameShape(QFrame.StyledPanel)
        card.setStyleSheet(f"background-color: {color}; border-radius: 5px; padding: 10px;")
        
        # Делаем кликабельной (если это active_breaks или over_limit)
        if card_id in ['active_breaks', 'over_limit']:
            card.setCursor(Qt.PointingHandCursor)
            card.mousePressEvent = lambda event, cid=card_id: self._on_dashboard_card_click(cid)
        
        layout = QVBoxLayout(card)
        
        # Заголовок
        title_label = QLabel(title)
        title_label.setStyleSheet("color: white; font-weight: bold;")
        layout.addWidget(title_label)
        
        # Значение
        value_label = QLabel(value)
        value_label.setStyleSheet("color: white; font-size: 24px; font-weight: bold;")
        value_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(value_label)
        
        # Сохраняем ссылку на label для обновления
        setattr(self, f"dashboard_{card_id}_label", value_label)
        
        return card
    
    def _build_filters(self):
        """Создаёт панель фильтров"""
        group = QGroupBox("🔍 Фильтры")
        layout = QHBoxLayout()
        
        # Период
        layout.addWidget(QLabel("Период:"))
        
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addDays(-7))
        layout.addWidget(self.date_from)
        
        layout.addWidget(QLabel("—"))
        
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        layout.addWidget(self.date_to)
        
        # Сотрудник
        layout.addWidget(QLabel("Сотрудник:"))
        self.filter_email = QLineEdit()
        self.filter_email.setPlaceholderText("Email или оставьте пустым")
        self.filter_email.setMinimumWidth(200)
        layout.addWidget(self.filter_email)
        
        # Группа
        layout.addWidget(QLabel("Группа:"))
        self.filter_group = QComboBox()
        self.filter_group.addItem("Все группы", None)
        # Загружаем группы
        try:
            ws = self.break_mgr.sheets.get_worksheet("Groups")
            groups_data = self.break_mgr.sheets._read_table(ws)
            for group_row in groups_data:
                group_name = group_row.get("Group", "")
                if group_name:
                    self.filter_group.addItem(group_name, group_name)
        except Exception as e:
            logger.warning(f"Failed to load groups: {e}")
        layout.addWidget(self.filter_group)
        
        # Тип нарушения
        layout.addWidget(QLabel("Тип нарушения:"))
        self.filter_violation_type = QComboBox()
        self.filter_violation_type.addItem("Все типы", None)
        self.filter_violation_type.addItem("Вне окна", "OUT_OF_WINDOW")
        self.filter_violation_type.addItem("Превышен лимит", "OVER_LIMIT")
        self.filter_violation_type.addItem("Превышено количество", "QUOTA_EXCEEDED")
        layout.addWidget(self.filter_violation_type)
        
        # Кнопка применить
        btn_apply = QPushButton("Применить")
        btn_apply.clicked.connect(self.apply_filters)
        layout.addWidget(btn_apply)
        
        layout.addStretch()
        
        group.setLayout(layout)
        return group
    
    def _build_stats(self):
        """Создаёт панель статистики"""
        group = QGroupBox("📊 Статистика")
        layout = QHBoxLayout()
        
        # Всего нарушений
        self.stat_total = QLabel("Всего: 0")
        self.stat_total.setStyleSheet("font-size: 14px; font-weight: bold;")
        layout.addWidget(self.stat_total)
        
        layout.addWidget(QLabel("|"))
        
        # По типам
        self.stat_out_of_window = QLabel("Вне окна: 0")
        layout.addWidget(self.stat_out_of_window)
        
        self.stat_over_limit = QLabel("Превышен лимит: 0")
        self.stat_over_limit.setStyleSheet("color: #e74c3c;")
        layout.addWidget(self.stat_over_limit)
        
        self.stat_quota_exceeded = QLabel("Превышено количество: 0")
        self.stat_quota_exceeded.setStyleSheet("color: #e74c3c;")
        layout.addWidget(self.stat_quota_exceeded)
        
        layout.addStretch()
        
        group.setLayout(layout)
        return group
    
    def _build_reports_tab(self):
        """Вкладка отчётов и графиков"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Заголовок
        title = QLabel("📈 ОТЧЁТЫ И АНАЛИТИКА")
        title.setStyleSheet("font-size: 16px; font-weight: bold;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Тип отчёта
        report_type_layout = QHBoxLayout()
        report_type_layout.addWidget(QLabel("Тип отчёта:"))
        
        self.report_type = QComboBox()
        self.report_type.addItem("Сводный отчёт за период", "summary")
        self.report_type.addItem("Сравнение групп", "groups")
        self.report_type.addItem("Топ нарушителей", "top_violators")
        self.report_type.addItem("Динамика по дням", "dynamics")
        self.report_type.addItem("Детальный отчёт по сотруднику", "employee_detail")
        report_type_layout.addWidget(self.report_type)
        
        report_type_layout.addStretch()
        layout.addLayout(report_type_layout)
        
        # Параметры отчёта
        params_group = QGroupBox("Параметры")
        params_layout = QVBoxLayout()
        
        # Период
        period_layout = QHBoxLayout()
        period_layout.addWidget(QLabel("Период:"))
        
        self.report_date_from = QDateEdit()
        self.report_date_from.setCalendarPopup(True)
        self.report_date_from.setDate(QDate.currentDate().addMonths(-1))
        period_layout.addWidget(self.report_date_from)
        
        period_layout.addWidget(QLabel("—"))
        
        self.report_date_to = QDateEdit()
        self.report_date_to.setCalendarPopup(True)
        self.report_date_to.setDate(QDate.currentDate())
        period_layout.addWidget(self.report_date_to)
        
        period_layout.addStretch()
        params_layout.addLayout(period_layout)
        
        params_group.setLayout(params_layout)
        layout.addWidget(params_group)
        
        # Кнопки генерации
        buttons_layout = QHBoxLayout()
        
        btn_generate = QPushButton("📊 Сгенерировать отчёт")
        btn_generate.clicked.connect(self.generate_report)
        buttons_layout.addWidget(btn_generate)
        
        btn_export_report = QPushButton("📥 Экспорт отчёта в Excel")
        btn_export_report.clicked.connect(self.export_report_to_excel)
        buttons_layout.addWidget(btn_export_report)
        
        buttons_layout.addStretch()
        layout.addLayout(buttons_layout)
        
        # Область для отображения отчёта
        self.report_display = QLabel("Выберите тип отчёта и нажмите 'Сгенерировать'")
        self.report_display.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        self.report_display.setWordWrap(True)
        self.report_display.setStyleSheet("background-color: white; padding: 10px; border: 1px solid #ccc;")
        self.report_display.setMinimumHeight(400)
        layout.addWidget(self.report_display)
        
        layout.addStretch()
        
        return widget
    
    # =================== ОБРАБОТЧИКИ ===================
    
    def refresh_dashboard(self):
        """Обновляет Dashboard"""
        try:
            # Получаем активные перерывы
            active_breaks = self._get_active_breaks()
            self.dashboard_active_breaks_data = active_breaks  # Сохраняем для клика
            self.dashboard_active_breaks_label.setText(f"{len(active_breaks)} человек")
            
            # Нарушений сегодня
            today = date.today().isoformat()
            violations = self.break_mgr.get_violations_report(
                date_from=today,
                date_to=today
            )
            self.dashboard_today_violations_label.setText(str(len(violations)))
            
            # Кто превышает лимит (из АКТИВНЫХ перерывов)
            over_limit_breaks = [b for b in active_breaks if b.get('is_over_limit', False)]
            over_limit_emails = set(b.get('Email') for b in over_limit_breaks if b.get('Email'))
            self.dashboard_over_limit_data = over_limit_breaks  # Сохраняем для клика
            self.dashboard_over_limit_label.setText(f"{len(over_limit_emails)} человек")
            
            # Топ нарушитель
            top_violator = self._get_top_violator(violations)
            if top_violator:
                self.dashboard_top_violator_label.setText(
                    f"{top_violator['email']}\n({top_violator['count']} нарушений)"
                )
            else:
                self.dashboard_top_violator_label.setText("Нет данных")
            
        except Exception as e:
            logger.error(f"Error refreshing dashboard: {e}")
    
    def apply_filters(self):
        """Применяет фильтры и обновляет таблицу"""
        try:
            # Получаем параметры фильтров
            date_from = self.date_from.date().toString("yyyy-MM-dd")
            date_to = self.date_to.date().toString("yyyy-MM-dd")
            email = self.filter_email.text().strip() or None
            violation_type = self.filter_violation_type.currentData()
            selected_group = self.filter_group.currentData()  # Получаем выбранную группу
            
            # Получаем данные
            violations = self.break_mgr.get_violations_report(
                email=email,
                date_from=date_from,
                date_to=date_to,
                violation_type=violation_type
            )
            
            # Фильтр по группе (если выбрана)
            if selected_group:
                # Получаем список email пользователей из группы
                try:
                    ws_users = self.break_mgr.sheets.get_worksheet("Users")
                    users = self.break_mgr.sheets._read_table(ws_users)
                    group_emails = [u.get("Email", "").lower() for u in users 
                                   if u.get("Group", "") == selected_group]
                    
                    # Фильтруем нарушения по email из группы
                    violations = [v for v in violations 
                                 if v.get("Email", "").lower() in group_emails]
                except Exception as e:
                    logger.warning(f"Failed to filter by group: {e}")
            
            self.current_violations = violations
            
            # Обновляем таблицу
            self._populate_violations_table(violations)
            
            # Обновляем статистику
            self._update_statistics(violations)
            
        except Exception as e:
            logger.error(f"Error applying filters: {e}")
            QMessageBox.warning(self, "Ошибка", f"Ошибка при загрузке данных: {e}")
    
    def export_to_excel(self):
        """Экспортирует данные в Excel"""
        if not self.current_violations:
            QMessageBox.information(self, "Информация", "Нет данных для экспорта")
            return
        
        # Диалог сохранения
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить отчёт",
            f"violations_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            self._export_violations_to_excel(self.current_violations, filename)
            QMessageBox.information(self, "Успех", f"Отчёт сохранён:\n{filename}")
        except Exception as e:
            logger.error(f"Export error: {e}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта: {e}")
    
    def generate_report(self):
        """Генерирует выбранный отчёт"""
        report_type = self.report_type.currentData()
        date_from = self.report_date_from.date().toString("yyyy-MM-dd")
        date_to = self.report_date_to.date().toString("yyyy-MM-dd")
        
        try:
            if report_type == "summary":
                report_text = self._generate_summary_report(date_from, date_to)
            elif report_type == "groups":
                report_text = self._generate_groups_comparison(date_from, date_to)
            elif report_type == "top_violators":
                report_text = self._generate_top_violators(date_from, date_to)
            elif report_type == "dynamics":
                report_text = self._generate_dynamics_report(date_from, date_to)
            elif report_type == "employee_detail":
                report_text = self._generate_employee_detail(date_from, date_to)
            else:
                report_text = "Неизвестный тип отчёта"
            
            self.report_display.setText(report_text)
            
        except Exception as e:
            logger.error(f"Error generating report: {e}")
            self.report_display.setText(f"Ошибка генерации отчёта: {e}")
    
    def export_report_to_excel(self):
        """Экспортирует текущий отчёт в Excel"""
        QMessageBox.information(self, "В разработке", "Экспорт отчётов в Excel будет добавлен в следующей версии")
    
    # =================== ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ ===================
    
    def _populate_violations_table(self, violations):
        """Заполняет таблицу нарушений"""
        self.violations_table.setRowCount(len(violations))
        
        for row, violation in enumerate(violations):
            # Дата/Время в московском времени (UTC+3)
            timestamp = violation.get("Timestamp", "")
            timestamp_formatted = format_datetime_moscow(timestamp) if timestamp else ""
            self.violations_table.setItem(row, 0, QTableWidgetItem(timestamp_formatted))
            
            # Сотрудник
            email = violation.get("Email", "")
            self.violations_table.setItem(row, 1, QTableWidgetItem(email))
            
            # Тип перерыва (извлекаем из Details если есть)
            details = violation.get("Details", "")
            break_type = "Перерыв" if "Перерыв" in details else "Обед" if "Обед" in details else "—"
            self.violations_table.setItem(row, 2, QTableWidgetItem(break_type))
            
            # Тип нарушения
            vtype = violation.get("ViolationType", "")
            vtype_text = {
                "OUT_OF_WINDOW": "Вне окна",
                "OVER_LIMIT": "Превышен лимит",
                "QUOTA_EXCEEDED": "Превышено количество"
            }.get(vtype, vtype)
            
            item = QTableWidgetItem(vtype_text)
            if vtype in ["OVER_LIMIT", "QUOTA_EXCEEDED"]:
                item.setForeground(QColor("#e74c3c"))
            self.violations_table.setItem(row, 3, item)
            
            # Детали
            self.violations_table.setItem(row, 4, QTableWidgetItem(details))
            
            # Критичность
            severity = "CRITICAL" if vtype in ["OVER_LIMIT", "QUOTA_EXCEEDED"] else "INFO"
            severity_text = "Критическое" if severity == "CRITICAL" else "Информация"
            item = QTableWidgetItem(severity_text)
            if severity == "CRITICAL":
                item.setForeground(QColor("#e74c3c"))
            self.violations_table.setItem(row, 5, item)
            
            # Статус
            status = violation.get("Status", "pending")
            status_text = {"pending": "Ожидает", "resolved": "Решено", "noted": "Отмечено"}.get(status, status)
            self.violations_table.setItem(row, 6, QTableWidgetItem(status_text))
    
    def _update_statistics(self, violations):
        """Обновляет статистику"""
        total = len(violations)
        out_of_window = len([v for v in violations if v.get("ViolationType") == "OUT_OF_WINDOW"])
        over_limit = len([v for v in violations if v.get("ViolationType") == "OVER_LIMIT"])
        quota_exceeded = len([v for v in violations if v.get("ViolationType") == "QUOTA_EXCEEDED"])
        
        self.stat_total.setText(f"Всего: {total}")
        self.stat_out_of_window.setText(f"Вне окна: {out_of_window}")
        self.stat_over_limit.setText(f"Превышен лимит: {over_limit}")
        self.stat_quota_exceeded.setText(f"Превышено количество: {quota_exceeded}")
    
    def _get_active_breaks(self):
        """Получает список активных перерывов"""
        try:
            return self.break_mgr.get_all_active_breaks()
        except Exception as e:
            logger.error(f"Failed to get active breaks: {e}")
            return []
    
    def _get_top_violator(self, violations):
        """Определяет топ нарушителя"""
        if not violations:
            return None
        
        # Подсчитываем нарушения по пользователям
        counts = {}
        for v in violations:
            email = v.get("Email", "")
            counts[email] = counts.get(email, 0) + 1
        
        if not counts:
            return None
        
        top_email = max(counts, key=counts.get)
        return {"email": top_email, "count": counts[top_email]}
    
    def _export_violations_to_excel(self, violations, filename):
        """Экспортирует нарушения в Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Нарушения"
            
            # Заголовки
            headers = ["Дата/Время", "Сотрудник", "Тип", "Тип нарушения", "Детали", "Критичность", "Статус"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Данные
            for row, violation in enumerate(violations, 2):
                timestamp = violation.get("Timestamp", "")
                timestamp_formatted = format_datetime_moscow(timestamp) if timestamp else ""
                ws.cell(row=row, column=1, value=timestamp_formatted)
                ws.cell(row=row, column=2, value=violation.get("Email", ""))
                
                details = violation.get("Details", "")
                break_type = "Перерыв" if "Перерыв" in details else "Обед" if "Обед" in details else "—"
                ws.cell(row=row, column=3, value=break_type)
                
                vtype = violation.get("ViolationType", "")
                vtype_text = {
                    "OUT_OF_WINDOW": "Вне окна",
                    "OVER_LIMIT": "Превышен лимит",
                    "QUOTA_EXCEEDED": "Превышено количество"
                }.get(vtype, vtype)
                ws.cell(row=row, column=4, value=vtype_text)
                
                ws.cell(row=row, column=5, value=details)
                
                severity = "CRITICAL" if vtype in ["OVER_LIMIT", "QUOTA_EXCEEDED"] else "INFO"
                ws.cell(row=row, column=6, value="Критическое" if severity == "CRITICAL" else "Информация")
                
                status = violation.get("Status", "pending")
                ws.cell(row=row, column=7, value={"pending": "Ожидает", "resolved": "Решено"}.get(status, status))
            
            # Автоширина колонок
            for col in range(1, 8):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
            
            wb.save(filename)
            
        except ImportError:
            raise Exception("Модуль openpyxl не установлен. Установите: pip install openpyxl --break-system-packages")
    
    def _generate_summary_report(self, date_from, date_to):
        """Генерирует сводный отчёт"""
        violations = self.break_mgr.get_violations_report(date_from=date_from, date_to=date_to)
        
        total = len(violations)
        
        # Защита от деления на ноль
        if total == 0:
            return f"""
СВОДНЫЙ ОТЧЁТ
Период: {date_from} — {date_to}

═══════════════════════════════════════

НЕТ ДАННЫХ

За выбранный период нарушения не зафиксированы.

═══════════════════════════════════════

Отчёт сгенерирован: {format_datetime_moscow(datetime.now())}
            """.strip()
        
        out_of_window = len([v for v in violations if v.get("ViolationType") == "OUT_OF_WINDOW"])
        over_limit = len([v for v in violations if v.get("ViolationType") == "OVER_LIMIT"])
        quota = len([v for v in violations if v.get("ViolationType") == "QUOTA_EXCEEDED"])
        
        # Уникальные нарушители
        violators = set(v.get("Email", "") for v in violations)
        
        report = f"""
СВОДНЫЙ ОТЧЁТ
Период: {date_from} — {date_to}

═══════════════════════════════════════

ОБЩАЯ СТАТИСТИКА:
  Всего нарушений: {total}
  Уникальных нарушителей: {len(violators)}

ПО ТИПАМ НАРУШЕНИЙ:
  • Вне временного окна: {out_of_window} ({out_of_window/total*100:.1f}%)
  • Превышен лимит времени: {over_limit} ({over_limit/total*100:.1f}%)
  • Превышено количество: {quota} ({quota/total*100:.1f}%)

КРИТИЧНОСТЬ:
  • Критические нарушения: {over_limit + quota}
  • Информационные: {out_of_window}

═══════════════════════════════════════

Отчёт сгенерирован: {format_datetime_moscow(datetime.now())}
        """
        
        return report.strip()
    
    def _generate_groups_comparison(self, date_from, date_to):
        """Сравнение групп"""
        return "Сравнение групп\n(В разработке - требуется информация о группах пользователей)"
    
    def _generate_top_violators(self, date_from, date_to):
        """Топ нарушителей"""
        violations = self.break_mgr.get_violations_report(date_from=date_from, date_to=date_to)
        
        # Подсчитываем по пользователям
        counts = {}
        for v in violations:
            email = v.get("Email", "")
            if email:
                counts[email] = counts.get(email, 0) + 1
        
        # Сортируем
        sorted_violators = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:10]
        
        report = f"ТОП-10 НАРУШИТЕЛЕЙ\nПериод: {date_from} — {date_to}\n\n"
        
        for idx, (email, count) in enumerate(sorted_violators, 1):
            report += f"{idx}. {email}: {count} нарушений\n"
        
        return report
    
    def _generate_dynamics_report(self, date_from, date_to):
        """Динамика по дням"""
        return "Динамика нарушений по дням\n(В разработке - требуется построение графика)"
    
    def _generate_employee_detail(self, date_from, date_to):
        """Детальный отчёт по сотруднику"""
        return "Детальный отчёт по сотруднику\n(Введите email в фильтрах и примените)"


    def _on_dashboard_card_click(self, card_id):
        """Обработчик клика на карточку Dashboard"""
        if card_id == 'active_breaks':
            self._show_active_breaks_dialog()
        elif card_id == 'over_limit':
            self._show_over_limit_dialog()
    
    def _show_active_breaks_dialog(self):
        """Показывает список пользователей в перерыве"""
        if not self.dashboard_active_breaks_data:
            QMessageBox.information(self, "Сейчас в перерыве", 
                                   "Никто не находится в перерыве")
            return
        
        # Создаём диалог
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, QPushButton, QHBoxLayout
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"👥 Сейчас в перерыве ({len(self.dashboard_active_breaks_data)} чел.)")
        dialog.resize(900, 500)
        
        layout = QVBoxLayout(dialog)
        
        # Таблица
        table = QTableWidget()
        table.setColumnCount(7)
        table.setHorizontalHeaderLabels([
            "Email", "Имя", "Тип", "Начало", "Длительность (мин)", "Статус", "Нарушение"
        ])
        table.horizontalHeader().setStretchLastSection(True)
        table.setRowCount(len(self.dashboard_active_breaks_data))
        
        # Заполняем
        for row, br in enumerate(self.dashboard_active_breaks_data):
            email = br.get('Email', 'N/A')
            name = br.get('Name', 'N/A')
            break_type = br.get('BreakType', 'N/A')
            start_time_raw = br.get('StartTime', 'N/A')
            # Форматируем время начала в московское (UTC+3)
            start = format_datetime_moscow(start_time_raw) if start_time_raw != 'N/A' else 'N/A'
            duration = br.get('Duration', 0)
            is_over = br.get('is_over_limit', False)
            is_violator = br.get('is_violator', False)
            violation_reason = br.get('violation_reason', '')
            
            table.setItem(row, 0, QTableWidgetItem(email))
            table.setItem(row, 1, QTableWidgetItem(name))
            table.setItem(row, 2, QTableWidgetItem(break_type))
            table.setItem(row, 3, QTableWidgetItem(start))
            table.setItem(row, 4, QTableWidgetItem(str(duration)))
            
            # Статус: норма или превышен лимит времени
            status_text = "✅ В норме"
            if is_over:
                status_text = "⚠️ Превышен лимит"
            status_item = QTableWidgetItem(status_text)
            if is_over:
                status_item.setBackground(QColor("#e74c3c"))
                status_item.setForeground(QColor("white"))
            table.setItem(row, 5, status_item)
            
            # Колонка нарушений
            violation_text = "✅ Норма" if not is_violator else f"❌ {violation_reason or 'Нарушитель'}"
            violation_item = QTableWidgetItem(violation_text)
            if is_violator:
                violation_item.setForeground(QColor(255, 140, 0))  # Оранжевый для нарушителей
                violation_item.setBackground(QColor(255, 250, 240))  # Светло-оранжевый фон
            table.setItem(row, 6, violation_item)
        
        table.resizeColumnsToContents()
        layout.addWidget(table)
        
        # Кнопка закрыть
        btn_close = QPushButton("Закрыть")
        btn_close.clicked.connect(dialog.close)
        layout.addWidget(btn_close)
        
        dialog.exec_()
    
    def _show_over_limit_dialog(self):
        """Показывает список пользователей с превышением лимита"""
        if not self.dashboard_over_limit_data:
            QMessageBox.information(self, "Превышают лимит", 
                                   "Никто не превышает лимит")
            return
        
        # Создаём диалог
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, QPushButton
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"⚠️ Превышают лимит ({len(self.dashboard_over_limit_data)} чел.)")
        dialog.resize(900, 500)
        
        layout = QVBoxLayout(dialog)
        
        # Таблица
        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels([
            "Email", "Имя", "Тип", "Начало", "Длительность (мин)", "Превышение (мин)"
        ])
        table.horizontalHeader().setStretchLastSection(True)
        table.setRowCount(len(self.dashboard_over_limit_data))
        
        # Заполняем
        for row, br in enumerate(self.dashboard_over_limit_data):
            email = br.get('Email', 'N/A')
            name = br.get('Name', 'N/A')
            break_type = br.get('BreakType', 'N/A')
            start_time_raw = br.get('StartTime', 'N/A')
            # Форматируем время начала в московское (UTC+3)
            start = format_datetime_moscow(start_time_raw) if start_time_raw != 'N/A' else 'N/A'
            duration = br.get('Duration', 0)
            
            # Вычисляем превышение
            limit = 15 if break_type == "Перерыв" else 60
            overage = duration - limit
            
            table.setItem(row, 0, QTableWidgetItem(email))
            table.setItem(row, 1, QTableWidgetItem(name))
            table.setItem(row, 2, QTableWidgetItem(break_type))
            table.setItem(row, 3, QTableWidgetItem(start))
            
            duration_item = QTableWidgetItem(str(duration))
            duration_item.setBackground(QColor("#e74c3c"))
            duration_item.setForeground(QColor("white"))
            table.setItem(row, 4, duration_item)
            
            overage_item = QTableWidgetItem(f"+{overage}")
            overage_item.setBackground(QColor("#c0392b"))
            overage_item.setForeground(QColor("white"))
            table.setItem(row, 5, overage_item)
        
        table.resizeColumnsToContents()
        layout.addWidget(table)
        
        # Кнопка закрыть
        btn_close = QPushButton("Закрыть")
        btn_close.clicked.connect(dialog.close)
        layout.addWidget(btn_close)
        
        dialog.exec_()


if __name__ == "__main__":
    from PyQt5.QtWidgets import QApplication
    import sys
    
    app = QApplication(sys.argv)
    
    # Mock manager для теста
    class MockBreakManager:
        def get_violations_report(self, **kwargs):
            return []
    
    widget = BreakAnalyticsTab(MockBreakManager())
    widget.setWindowTitle("Break Analytics Test")
    widget.resize(1200, 800)
    widget.show()
    

    sys.exit(app.exec_())